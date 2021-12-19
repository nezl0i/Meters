from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import openpyxl


def to_real_time(stamp):
    return datetime.fromtimestamp(int(stamp)).strftime("%d.%m.%Y")


def to_stamp_time(realtime):
    return int(datetime.timestamp(datetime.strptime(realtime, "%d.%m.%Y")))


print(to_real_time('1590576546'))
print(to_stamp_time('25.12.2021'))

# shtamp = 1590576546
#
# realtime = datetime.fromtimestamp(shtamp)   # time is 27-05-2020 13:05:06
# to_time = (datetime.fromtimestamp(shtamp)).strftime("%d.%m.%Y")    # format to str 27.05.2020 13:05:06
# to_real_time = datetime.strptime(to_time, "%d.%m.%Y")
# to_shtamp = int(datetime.timestamp(realtime))   # convert to shtamp int 1590576546

# print(to_time)
# print(to_shtamp)
# print(to_real_time)


# srcfile = openpyxl.load_workbook('test.xlsx', read_only=False, keep_vba=True)
#
# _sheet_name = srcfile['Разбивка по филиалам']
# _sheet_name['B2'] = 'write something'
# _sheet_name.cell(row=1, column=1).value = "Записал"
#
# srcfile.save('test.xlsx')

